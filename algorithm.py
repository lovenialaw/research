"""
algorithm.py

Find an (approximately) optimal gearbox disassembly sequence using:
  - Precedence constraints from `gearbox_kg.html`
  - Uncertainties from `gearbox.xlsx` (sheets: components, tools)

Goal:
  - remove a chosen node (default: C1 Base) while minimizing expected cost

Notes:
  - The connection graph encodes precedence: edge A -> B means B can only be removed after A.
  - We model stochastic "remove attempt success" using uncertainties:
      - Fastener nodes (F*) use rows in `tools` mapped by Sheet6 (fastener -> BOM part).
      - Component nodes (C*) use rows in `components` mapped by Sheet5 (component -> BOM part).
  - Failure model used here:
      - "retry until success" for the chosen node.
      - Tool wear accumulates on each fastener removal attempt, which makes the route/order matter.

This script provides:
  - Dijkstra-style baseline (no wear, deterministic expected cost per node) producing a candidate sequence.
  - Genetic Algorithm (GA) optimizing a priority-weight vector, evaluated by Monte Carlo simulation
    using wear + stochastic success probabilities.

Run examples:
  python algorithm.py --method dijkstra --goal C1
  python algorithm.py --method ga --goal C1 --pop 40 --gens 50 --sims 200
  python algorithm.py --method ga --dynamic-mc --sims 200 --sims-coarse 40 --fine-top-frac 0.1
"""

from __future__ import annotations

import argparse
import math
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional, Callable

_RESEARCH_ROOT = Path(__file__).resolve().parent
_DEFAULT_XLSX = str(_RESEARCH_ROOT / "gearbox.xlsx")
_DEFAULT_GRAPH = str(_RESEARCH_ROOT / "gearbox_kg.html")


def parse_graph_html_edges(html_path: str) -> Tuple[Set[str], Dict[str, Set[str]]]:
    """
    Returns:
      nodes: all node ids found in the edges array
      preds: mapping node -> set of predecessor nodes
    """
    with open(html_path, "r", encoding="utf-8") as f:
        txt = f.read()

    # Extract the edges array block inside: const edges = new vis.DataSet([ ... ]).map(...)
    # Then parse entries like: ['F1','F2']
    edges_block_start = txt.find("const edges")
    if edges_block_start < 0:
        raise RuntimeError("Could not find 'const edges' in graph html")

    edges_block_end = txt.find("].map", edges_block_start)
    if edges_block_end < 0:
        raise RuntimeError("Could not find edges array ending in graph html")

    block = txt[edges_block_start:edges_block_end]
    pair_pat = re.compile(r"\[\s*'([^']+)'\s*,\s*'([^']+)'\s*\]")
    pairs = pair_pat.findall(block)

    if not pairs:
        raise RuntimeError("No edges parsed from graph html; regex mismatch")

    preds: Dict[str, Set[str]] = {}
    nodes: Set[str] = set()
    for a, b in pairs:
        nodes.add(a)
        nodes.add(b)
        preds.setdefault(b, set()).add(a)
        preds.setdefault(a, set())

    # Ensure all nodes exist in preds map
    for n in nodes:
        preds.setdefault(n, set())

    return nodes, preds


def ancestors_of_goal(goal: str, preds: Dict[str, Set[str]]) -> Set[str]:
    """Return all nodes that must be removed before `goal` is eligible (graph ancestors)."""
    seen: Set[str] = set()
    stack = [goal]
    while stack:
        cur = stack.pop()
        for p in preds.get(cur, set()):
            if p not in seen:
                seen.add(p)
                stack.append(p)
    seen.add(goal)
    return seen


def topo_available(removed: Set[str], preds: Dict[str, Set[str]], remaining: Set[str]) -> List[str]:
    """Available nodes among `remaining` given that all predecessors are removed."""
    avail = []
    for v in remaining:
        if v in removed:
            continue
        if preds[v].issubset(removed):
            avail.append(v)
    return avail


@dataclass(frozen=True)
class FastenerUnc:
    tool_success_rate: float
    tool_wear: float
    bolt_seizure_probability: float
    bearing_jamming_probability: float
    removal_force_Nm: float
    task_code: Optional[str] = None
    tool_code: Optional[str] = None


@dataclass(frozen=True)
class ComponentUnc:
    temperature: float
    corrosion_level: float
    task_code: Optional[str] = None


def clamp01(x: float) -> float:
    return max(0.0, min(1.0, x))


def load_excel_uncertainties(
    xlsx_path: str,
    verbose: bool = False,
) -> Tuple[
    Dict[int, ComponentUnc],
    Dict[int, FastenerUnc],
    Dict[str, int],
    Dict[str, int],
    Dict[str, float],
]:
    """
    Returns:
      components_unc_by_part: BOM part code -> ComponentUnc
      tools_unc_by_part: BOM part code -> FastenerUnc
      component_node_to_part: graph node id (e.g. 'C2SX') -> BOM part code
      fastener_node_to_part: graph node id (e.g. 'F3DX') -> BOM part code
    """
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to parse gearbox.xlsx. Install it:\n"
            "  pip install openpyxl\n"
            f"Original error: {e}"
        )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def _norm_header(x) -> str:
        return re.sub(r"\s+", " ", str(x).strip().lower()) if x is not None else ""

    def _build_col_map(ws, header_row: int) -> Dict[str, int]:
        headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        m: Dict[str, int] = {}
        for idx, h in enumerate(headers, start=1):
            k = _norm_header(h)
            if not k:
                continue
            m.setdefault(k, idx)
        return m

    def _find_task_table_row(ws) -> Optional[int]:
        for r in range(1, min(500, ws.max_row + 1)):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if a is None or b is None:
                continue
            if "task code" in str(a).lower() and "difficulty" in str(b).lower():
                return r
        return None

    # Sheet components: two-row header, includes Task column now
    ws_comp = wb["components"]
    comp_row1 = _build_col_map(ws_comp, header_row=1)
    comp_row2 = _build_col_map(ws_comp, header_row=2)
    components_unc_by_part: Dict[int, ComponentUnc] = {}
    part_col = comp_row1.get("part number", 1)
    task_col = comp_row1.get("task", 4)
    temp_col = comp_row2.get("temperature", 5)
    corr_col = comp_row2.get("corrosion level", 6)

    task_table_row = _find_task_table_row(ws_comp)
    stop_row = (task_table_row - 1) if task_table_row else ws_comp.max_row

    for r in range(3, stop_row + 1):
        part_raw = ws_comp.cell(r, part_col).value
        if part_raw is None:
            continue
        try:
            part = int(float(part_raw))
        except Exception:
            continue
        temp = ws_comp.cell(r, temp_col).value
        corr = ws_comp.cell(r, corr_col).value
        task = ws_comp.cell(r, task_col).value if task_col else None
        if temp is None or corr is None:
            continue
        task_code = str(task).strip() if task is not None and str(task).strip() else None
        components_unc_by_part[part] = ComponentUnc(temperature=float(temp), corrosion_level=float(corr), task_code=task_code)

    # Task code -> difficulty table
    task_difficulty_by_code: Dict[str, float] = {}
    if task_table_row:
        for r in range(task_table_row + 1, min(task_table_row + 60, ws_comp.max_row + 1)):
            code_cell = ws_comp.cell(r, 1).value
            diff_cell = ws_comp.cell(r, 2).value
            if code_cell is None or diff_cell is None:
                continue
            raw = str(code_cell).strip()
            if not raw:
                continue
            code = raw.split("-", 1)[0].strip()
            try:
                diff = float(diff_cell)
            except Exception:
                continue
            if code:
                task_difficulty_by_code[code] = diff

    # Sheet fasteners: renamed from "tools", includes Task column
    ws_tools = wb["fasteners"]
    fast_row1 = _build_col_map(ws_tools, header_row=1)
    fast_row2 = _build_col_map(ws_tools, header_row=2)
    tools_unc_by_part: Dict[int, FastenerUnc] = {}
    part_col_f = fast_row1.get("part number", 1)
    task_col_f = fast_row1.get("task", 4)
    tool_code_col_f = fast_row1.get("tool code") or fast_row1.get("tool") or fast_row1.get("tool/alttools")

    success_col = fast_row2.get("fasteners success rate") or fast_row2.get("tool success rate") or 5
    wear_col = fast_row2.get("fasteners wear") or fast_row2.get("tool wear") or 6
    bolt_col = fast_row2.get("bolt seizure probability") or 9
    jam_col = fast_row2.get("bearing jamming probability") or 10
    force_col = fast_row2.get("removal force (nm)") or 11

    for r in range(3, ws_tools.max_row + 1):
        part_raw = ws_tools.cell(r, part_col_f).value
        if part_raw is None:
            continue
        try:
            part = int(float(part_raw))
        except Exception:
            continue
        tool_success_rate = ws_tools.cell(r, success_col).value
        tool_wear = ws_tools.cell(r, wear_col).value
        bolt_seizure_prob = ws_tools.cell(r, bolt_col).value
        bearing_jam_prob = ws_tools.cell(r, jam_col).value
        removal_force_Nm = ws_tools.cell(r, force_col).value
        task = ws_tools.cell(r, task_col_f).value if task_col_f else None
        tool_code = ws_tools.cell(r, tool_code_col_f).value if tool_code_col_f else None

        if tool_success_rate is None or tool_wear is None:
            continue
        task_code = str(task).strip() if task is not None and str(task).strip() else None
        tool_code_s = str(tool_code).strip() if tool_code is not None and str(tool_code).strip() else None
        tools_unc_by_part[part] = FastenerUnc(
            tool_success_rate=float(tool_success_rate),
            tool_wear=float(tool_wear),
            bolt_seizure_probability=float(bolt_seizure_prob) if bolt_seizure_prob is not None else 0.0,
            bearing_jamming_probability=float(bearing_jam_prob) if bearing_jam_prob is not None else 0.0,
            removal_force_Nm=float(removal_force_Nm) if removal_force_Nm is not None else 0.0,
            task_code=task_code,
            tool_code=tool_code_s,
        )

    # Sheet5: component node -> BOM part
    ws5 = wb["Sheet5"]
    component_node_to_part: Dict[str, int] = {}
    for row in ws5.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        node = str(row[0])
        bom = row[1]
        if bom is None:
            continue
        component_node_to_part[node] = int(float(bom))

    # Sheet6: fastener node -> BOM part
    ws6 = wb["Sheet6"]
    fastener_node_to_part: Dict[str, int] = {}
    for row in ws6.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        node = str(row[0])
        bom = row[1]
        if bom is None:
            continue
        fastener_node_to_part[node] = int(float(bom))

    if verbose:
        print("Loaded:")
        print(f"  component uncertainties: {len(components_unc_by_part)} part codes")
        print(f"  tool uncertainties: {len(tools_unc_by_part)} part codes")
        print(f"  component node mapping: {len(component_node_to_part)} nodes")
        print(f"  fastener node mapping: {len(fastener_node_to_part)} nodes")
        print(f"  task difficulty codes: {len(task_difficulty_by_code)} codes -> {task_difficulty_by_code}")

    return components_unc_by_part, tools_unc_by_part, component_node_to_part, fastener_node_to_part, task_difficulty_by_code


def build_component_prob_fn(components_unc_by_part: Dict[int, ComponentUnc], a: float, b: float):
    """
    Returns p(C) based on component temperature + corrosion level.
    Simple model:
      p = exp(-a * corrosion_level) * exp(-b * temp_norm)
    where temp_norm is in [0,1].
    """
    temps = [u.temperature for u in components_unc_by_part.values()]
    if not temps:
        raise RuntimeError("No component uncertainties loaded (components sheet empty?)")
    tmin = min(temps)
    tmax = max(temps)
    denom = (tmax - tmin) if (tmax - tmin) != 0 else 1.0

    def p_component(part_code: int) -> float:
        u = components_unc_by_part.get(part_code)
        if u is None:
            return 0.5
        temp_norm = (u.temperature - tmin) / denom
        p = math.exp(-a * u.corrosion_level) * math.exp(-b * temp_norm)
        return clamp01(p)

    return p_component


def build_fastener_prob_fn(
    tools_unc_by_part: Dict[int, FastenerUnc],
    wear_k: float,
):
    """
    Returns p(F) based on:
      p0 = tool_success_rate * (1 - bolt_seizure_probability) * (1 - bearing_jamming_probability)
      p = p0 * exp(-wear_k * wear_state)
    """
    def p_fastener(part_code: int, wear_state: float) -> float:
        u = tools_unc_by_part.get(part_code)
        if u is None:
            return 0.5
        p0 = u.tool_success_rate * (1.0 - clamp01(u.bolt_seizure_probability)) * (1.0 - clamp01(u.bearing_jamming_probability))
        p = p0 * math.exp(-wear_k * max(0.0, wear_state))
        return clamp01(p)

    return p_fastener


def greedy_deterministic_sequence(
    nodes: Set[str],
    preds: Dict[str, Set[str]],
    goal: str,
    p_component_fn,
    p_fastener_fn,
    node_to_part: Dict[str, int],
    max_nodes: Optional[int] = None,
):
    """
    Deterministic baseline:
      choose among available nodes the one with minimum expected removal attempts (1/p),
      where p ignores wear (wear_state=0).
    """
    remaining = set(nodes)
    removed: Set[str] = set()
    seq: List[str] = []

    wear_state = 0.0  # ignored in p here (baseline uses wear-free)
    for _ in range(len(nodes) + 5):
        if goal in removed:
            break
        avail = topo_available(removed, preds, remaining)
        if not avail:
            break
        # optionally limit to ancestor set size
        if max_nodes is not None and len(removed) >= max_nodes:
            break
        def exp_attempts(v: str) -> float:
            part = node_to_part.get(v)
            if part is None:
                return 10.0
            if v.startswith("F"):
                p = p_fastener_fn(part, wear_state=0.0)
            else:
                p = p_component_fn(part)
            # Avoid division by zero
            p = max(1e-6, p)
            return 1.0 / p

        v_best = min(avail, key=exp_attempts)
        removed.add(v_best)
        seq.append(v_best)
    return seq, removed


def simulate_disassembly(
    preds: Dict[str, Set[str]],
    nodes_to_consider: Set[str],
    goal: str,
    component_node_to_part: Dict[str, int],
    fastener_node_to_part: Dict[str, int],
    p_component_fn,
    p_fastener_fn,
    wear_k: float,
    tool_wear_state: float = 0.0,
    rng: Optional[random.Random] = None,
    max_successful_removals: int = 1000,
    force_scale: float = 0.0,
) -> Tuple[List[str], int, float]:
    """
    Monte Carlo simulation using "retry until success":
      - choose next node by external policy (here: sequence is generated externally)
      - but for this function, we only compute stochastic removal when a node is attempted,
        and then the caller controls which node to attempt next by maintaining removed set.

    For simplicity, this function assumes the caller uses a fixed removal order `seq_order`.
    To keep it standalone, we implement policy "attempt nodes in the order you pass".
    """
    raise NotImplementedError("Use simulate_with_policy()")


def simulate_with_policy(
    preds: Dict[str, Set[str]],
    nodes_to_consider: Set[str],
    goal: str,
    component_node_to_part: Dict[str, int],
    fastener_node_to_part: Dict[str, int],
    p_component_fn,
    p_fastener_fn,
    tools_unc_by_part: Dict[int, FastenerUnc],
    policy_choose_next,
    rng: random.Random,
    force_scale: float = 0.0,
    wear_k: float = 1.0,
    max_successful_removals: int = 1000,
    max_attempts_total: int = 20000,
    task_difficulty_by_code: Optional[Dict[str, float]] = None,
    # Shared-tool heat model (applies only to fasteners; based on tool_code)
    heat_p: float = 0.20,
    heat_f: float = 0.15,
    heat_cool: int = 1,
    heat_max: int = 3,
    difficulty_cost_k: float = 0.10,
    difficulty_p_k: float = 0.05,
) -> Tuple[List[str], int, float]:
    """
    Run stochastic disassembly simulation until `goal` is removed.

    policy_choose_next(available_nodes, removed_set, wear_state)-> node
    """
    removed: Set[str] = set()
    wear_state = 0.0
    seq: List[str] = []
    attempts_total = 0
    cost_total = 0.0

    remaining = set(nodes_to_consider)
    last_tool_code: Optional[str] = None
    heat_level: int = 0

    while goal not in removed:
        if len(seq) > max_successful_removals:
            break
        if attempts_total >= max_attempts_total:
            # Abort runaway retries (e.g., extremely low p under compounded penalties).
            return seq, attempts_total, cost_total + 1e9
        available = topo_available(removed, preds, remaining)
        if not available:
            # stuck due to inconsistent precedence (should not happen)
            break
        v = policy_choose_next(available, removed, wear_state)
        if v not in available:
            # Invalid policy; fall back to random available to avoid crash.
            v = rng.choice(available)

        part_code = None
        is_fastener = v.startswith("F")
        if is_fastener:
            part_code = fastener_node_to_part.get(v)
        else:
            part_code = component_node_to_part.get(v)

        if part_code is None:
            # Unknown uncertainty mapping -> assume moderate difficulty
            part_code = -1

        # Retry until success on the chosen node
        while v not in removed:
            if attempts_total >= max_attempts_total:
                return seq, attempts_total, cost_total + 1e9
            is_fastener = v.startswith("F")
            if is_fastener:
                p = p_fastener_fn(part_code, wear_state)
                # Attempt cost: 1 + scaled removal force
                u = tools_unc_by_part.get(part_code)
                force_nm = u.removal_force_Nm if u else 0.0
                wear_increment = u.tool_wear if u else 0.0

                # Task difficulty effects (optional)
                diff = 0.0
                if task_difficulty_by_code and u and u.task_code and u.task_code in task_difficulty_by_code:
                    diff = float(task_difficulty_by_code[u.task_code])

                # Shared-tool heat: if same tool as last fastener, apply penalty
                # Use task_code as shared-resource identity (requested).
                # This means repeated tasks (e.g. many unscrew operations) can accumulate "heat".
                tool_code = (u.task_code if (u and u.task_code) else str(part_code))
                heat_pen = heat_level if (last_tool_code is not None and tool_code == last_tool_code) else 0
                p = p * math.exp(-heat_p * heat_pen) * math.exp(-difficulty_p_k * diff)

                attempt_cost = 1.0 + (force_scale * force_nm) * (1.0 + heat_f * heat_pen) * (1.0 + difficulty_cost_k * diff)
            else:
                p = p_component_fn(part_code)
                attempt_cost = 1.0
                wear_increment = 0.0

            p = max(1e-6, p)
            attempts_total += 1
            cost_total += attempt_cost

            # Update wear on every attempt (including success)
            wear_state += wear_increment

            roll = rng.random()
            if roll < p:
                removed.add(v)
                seq.append(v)
                # Update heat on successful removals (order-level effect)
                if v.startswith("F"):
                    u = tools_unc_by_part.get(part_code)
                    tool_code = (u.task_code if (u and u.task_code) else str(part_code))
                    if last_tool_code is not None and tool_code == last_tool_code:
                        heat_level = min(heat_max, heat_level + 1)
                    else:
                        heat_level = 1
                    last_tool_code = tool_code
                else:
                    heat_level = max(0, heat_level - heat_cool)
            else:
                # failure: state unchanged except wear/tool condition updated above
                pass

    return seq, attempts_total, cost_total


def evaluate_sequence_by_priority_weights(
    weights_by_node: Dict[str, float],
    preds: Dict[str, Set[str]],
    nodes_to_consider: Set[str],
    goal: str,
    component_node_to_part: Dict[str, int],
    fastener_node_to_part: Dict[str, int],
    p_component_fn,
    p_fastener_fn,
    tools_unc_by_part: Dict[int, FastenerUnc],
    rng: random.Random,
    sims: int,
    force_scale: float,
    task_difficulty_by_code: Optional[Dict[str, float]] = None,
) -> Tuple[float, float]:
    """
    Evaluate expected cost via Monte Carlo:
      policy: choose available node with highest weight.
    Returns:
      (mean_cost, mean_attempts)
    """

    def policy_choose_next(available, removed, wear_state):
        best = None
        best_w = -1e18
        for v in available:
            w = weights_by_node.get(v, 0.0)
            if w > best_w:
                best_w = w
                best = v
        return best

    costs = []
    attempts = []
    for _ in range(sims):
        sim_rng = random.Random(rng.randint(0, 10**9))
        _, att, cost = simulate_with_policy(
            preds=preds,
            nodes_to_consider=nodes_to_consider,
            goal=goal,
            component_node_to_part=component_node_to_part,
            fastener_node_to_part=fastener_node_to_part,
            p_component_fn=p_component_fn,
            p_fastener_fn=p_fastener_fn,
            tools_unc_by_part=tools_unc_by_part,
            policy_choose_next=policy_choose_next,
            rng=sim_rng,
            force_scale=force_scale,
            task_difficulty_by_code=task_difficulty_by_code,
        )
        costs.append(cost)
        attempts.append(att)

    mean_cost = sum(costs) / len(costs) if costs else 1e18
    mean_attempts = sum(attempts) / len(attempts) if attempts else 1e18
    return mean_cost, mean_attempts


def _eval_weights(
    weights_by_node: Dict[str, float],
    preds: Dict[str, Set[str]],
    nodes_to_consider: Set[str],
    goal: str,
    component_node_to_part: Dict[str, int],
    fastener_node_to_part: Dict[str, int],
    p_component_fn,
    p_fastener_fn,
    tools_unc_by_part: Dict[int, FastenerUnc],
    rng: random.Random,
    sims: int,
    force_scale: float,
    task_difficulty_by_code: Optional[Dict[str, float]] = None,
) -> float:
    mean_cost, _mean_attempts = evaluate_sequence_by_priority_weights(
        weights_by_node=weights_by_node,
        preds=preds,
        nodes_to_consider=nodes_to_consider,
        goal=goal,
        component_node_to_part=component_node_to_part,
        fastener_node_to_part=fastener_node_to_part,
        p_component_fn=p_component_fn,
        p_fastener_fn=p_fastener_fn,
        tools_unc_by_part=tools_unc_by_part,
        rng=rng,
        sims=sims,
        force_scale=force_scale,
        task_difficulty_by_code=task_difficulty_by_code,
    )
    return mean_cost


def genetic_algorithm_optimize(
    preds: Dict[str, Set[str]],
    nodes_to_consider: Set[str],
    goal: str,
    component_node_to_part: Dict[str, int],
    fastener_node_to_part: Dict[str, int],
    p_component_fn,
    p_fastener_fn,
    tools_unc_by_part: Dict[int, FastenerUnc],
    rng: random.Random,
    pop: int,
    gens: int,
    sims: int,
    force_scale: float,
    mutation_sigma: float,
    elite_frac: float,
    dynamic_mc: bool = False,
    sims_coarse: Optional[int] = None,
    fine_top_frac: float = 0.1,
    task_difficulty_by_code: Optional[Dict[str, float]] = None,
) -> Tuple[Dict[str, float], float]:
    """
    Chromosome:
      real-valued weight per node in nodes_to_consider

    At each step:
      choose available node with maximum weight.
    """
    nodes = sorted(nodes_to_consider)

    def make_individual():
        # initialize near 0 so choices are not too biased
        return {v: rng.uniform(-1.0, 1.0) for v in nodes}

    def crossover(a: Dict[str, float], b: Dict[str, float]) -> Dict[str, float]:
        # blend crossover
        alpha = rng.uniform(0.2, 0.8)
        return {v: alpha * a[v] + (1.0 - alpha) * b[v] for v in nodes}

    def mutate(ind: Dict[str, float]) -> Dict[str, float]:
        # gaussian mutation
        out = dict(ind)
        for v in nodes:
            if rng.random() < 0.2:  # mutate subset
                out[v] += rng.gauss(0.0, mutation_sigma)
        return out

    # initial population
    population: List[Dict[str, float]] = [make_individual() for _ in range(pop)]

    best_weights = None
    best_cost = float("inf")

    elite_n = max(1, int(pop * elite_frac))
    coarse_n = sims_coarse if sims_coarse is not None else max(5, sims // 5)
    coarse_n = max(1, min(coarse_n, sims))

    for g in range(gens):
        t0 = time.time()

        scored: List[Tuple[float, Dict[str, float]]] = []

        if dynamic_mc and coarse_n < sims:
            verify_n = max(elite_n, int(math.ceil(pop * fine_top_frac)))
            verify_n = min(verify_n, pop)

            coarse_costs: List[float] = []
            for ind in population:
                coarse_costs.append(
                    _eval_weights(
                        ind,
                        preds,
                        nodes_to_consider,
                        goal,
                        component_node_to_part,
                        fastener_node_to_part,
                        p_component_fn,
                        p_fastener_fn,
                        tools_unc_by_part,
                        rng,
                        coarse_n,
                        force_scale,
                        task_difficulty_by_code=task_difficulty_by_code,
                    )
                )

            order = sorted(range(pop), key=lambda i: coarse_costs[i])
            refined = set(order[:verify_n])
            final_costs = list(coarse_costs)
            for i in refined:
                final_costs[i] = _eval_weights(
                    population[i],
                    preds,
                    nodes_to_consider,
                    goal,
                    component_node_to_part,
                    fastener_node_to_part,
                    p_component_fn,
                    p_fastener_fn,
                    tools_unc_by_part,
                    rng,
                    sims,
                    force_scale,
                    task_difficulty_by_code=task_difficulty_by_code,
                )

            for i in range(pop):
                c = final_costs[i]
                ind = population[i]
                scored.append((c, ind))
                if c < best_cost:
                    best_cost = c
                    best_weights = dict(ind)
        else:
            for ind in population:
                mean_cost = _eval_weights(
                    ind,
                    preds,
                    nodes_to_consider,
                    goal,
                    component_node_to_part,
                    fastener_node_to_part,
                    p_component_fn,
                    p_fastener_fn,
                    tools_unc_by_part,
                    rng,
                    sims,
                    force_scale,
                    task_difficulty_by_code=task_difficulty_by_code,
                )
                scored.append((mean_cost, ind))

                if mean_cost < best_cost:
                    best_cost = mean_cost
                    best_weights = dict(ind)

        scored.sort(key=lambda x: x[0])
        elites = [ind for (_c, ind) in scored[:elite_n]]

        # print generation summary
        if g % 1 == 0:
            print(f"[GA] gen {g+1}/{gens} best_mean_cost={scored[0][0]:.4f} global_best={best_cost:.4f} time={time.time()-t0:.2f}s")

        # breed next generation
        next_pop = list(elites)
        while len(next_pop) < pop:
            a = rng.choice(elites)
            b = rng.choice(elites)
            child = crossover(a, b)
            child = mutate(child)
            next_pop.append(child)

        population = next_pop

    return best_weights or {}, best_cost


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=_DEFAULT_XLSX)
    parser.add_argument("--graph", default=_DEFAULT_GRAPH)
    parser.add_argument("--goal", default="C1")
    parser.add_argument("--method", choices=["dijkstra", "ga"], default="ga")

    # probability model parameters
    parser.add_argument("--comp_a", type=float, default=0.25, help="component corrosion exponent")
    parser.add_argument("--comp_b", type=float, default=1.00, help="component temperature exponent")
    parser.add_argument("--wear_k", type=float, default=1.25, help="tool wear exponent in fastener success prob")

    # simulation / cost
    parser.add_argument("--force_scale", type=float, default=0.002, help="time proxy per removal_force_Nm (fasteners only)")
    parser.add_argument("--sims", type=int, default=200, help="Monte Carlo sims for GA fitness (fine / full eval)")
    parser.add_argument("--seed", type=int, default=1234)

    # GA parameters
    parser.add_argument("--pop", type=int, default=30)
    parser.add_argument("--gens", type=int, default=25)
    parser.add_argument("--mutation_sigma", type=float, default=0.35)
    parser.add_argument("--elite_frac", type=float, default=0.25)
    parser.add_argument(
        "--dynamic-mc",
        action="store_true",
        help="Screen all individuals with fewer MC sims, then re-run full --sims on top fraction (see --sims-coarse, --fine-top-frac)",
    )
    parser.add_argument(
        "--sims-coarse",
        type=int,
        default=None,
        help="MC sims for screening when --dynamic-mc (default: max(5, --sims//5))",
    )
    parser.add_argument(
        "--fine-top-frac",
        type=float,
        default=0.1,
        help="With --dynamic-mc, re-verify best max(elites, ceil(pop*frac)) individuals with full --sims",
    )

    args = parser.parse_args()
    rng = random.Random(args.seed)

    nodes, preds = parse_graph_html_edges(args.graph)
    if args.goal not in nodes:
        raise RuntimeError(f"Goal node {args.goal} not found in graph html. Found nodes: {sorted(nodes)}")

    nodes_to_consider = ancestors_of_goal(args.goal, preds)
    # Baseline planning should only consider nodes that unlock goal
    print(f"[INFO] nodes in graph: {len(nodes)}")
    print(f"[INFO] ancestors needed for goal {args.goal}: {len(nodes_to_consider)} -> {sorted(nodes_to_consider)}")

    comps_unc, tools_unc, component_node_to_part, fastener_node_to_part, task_difficulty_by_code = load_excel_uncertainties(args.xlsx, verbose=True)
    p_component_fn = build_component_prob_fn(comps_unc, a=args.comp_a, b=args.comp_b)
    p_fastener_fn = build_fastener_prob_fn(tools_unc, wear_k=args.wear_k)

    def task_for_node(node_id: str) -> str:
        if node_id.startswith("F"):
            part = fastener_node_to_part.get(node_id)
            u = tools_unc.get(part) if part is not None else None
            return u.task_code if (u and u.task_code) else "?"
        part = component_node_to_part.get(node_id)
        u = comps_unc.get(part) if part is not None else None
        return u.task_code if (u and u.task_code) else "?"

    # Build unified node->part code mapping for baseline p0
    node_to_part: Dict[str, int] = {}
    node_to_part.update(component_node_to_part)
    node_to_part.update(fastener_node_to_part)

    if args.method == "dijkstra":
        seq, removed = greedy_deterministic_sequence(
            nodes=nodes_to_consider,
            preds=preds,
            goal=args.goal,
            p_component_fn=p_component_fn,
            p_fastener_fn=p_fastener_fn,
            node_to_part=node_to_part,
        )
        print("\n[Dijkstra-baseline] Sequence until goal:")
        print(" -> ".join(seq))
        print("[Dijkstra-baseline] Tasks:")
        print(" -> ".join(f"{v}({task_for_node(v)})" for v in seq))

        # Evaluate with Monte Carlo wear model to compare fairly
        # policy: follow greedy seq, but if next in seq isn't available we pick random available.
        seq_order = seq
        seq_idx = {v: i for i, v in enumerate(seq_order)}

        def policy_choose_next(avail, removed_set, wear_state):
            # choose among available with smallest index in greedy sequence
            best = None
            best_i = 10**9
            for v in avail:
                i = seq_idx.get(v, 10**9)
                if i < best_i:
                    best_i = i
                    best = v
            return best if best is not None else rng.choice(avail)

        # Run MC evaluation
        costs = []
        for _ in range(max(50, args.sims // 4)):
            sim_rng = random.Random(rng.randint(0, 10**9))
            _seq, _att, cost = simulate_with_policy(
                preds=preds,
                nodes_to_consider=nodes_to_consider,
                goal=args.goal,
                component_node_to_part=component_node_to_part,
                fastener_node_to_part=fastener_node_to_part,
                p_component_fn=p_component_fn,
                p_fastener_fn=p_fastener_fn,
                tools_unc_by_part=tools_unc,
                policy_choose_next=policy_choose_next,
                rng=sim_rng,
                force_scale=args.force_scale,
                wear_k=args.wear_k,
                task_difficulty_by_code=task_difficulty_by_code,
            )
            costs.append(cost)
        mean_cost = sum(costs) / len(costs)
        print(f"[Dijkstra-baseline] MC mean_cost={mean_cost:.4f} (n={len(costs)}) with wear model")

    elif args.method == "ga":
        best_weights, best_cost = genetic_algorithm_optimize(
            preds=preds,
            nodes_to_consider=nodes_to_consider,
            goal=args.goal,
            component_node_to_part=component_node_to_part,
            fastener_node_to_part=fastener_node_to_part,
            p_component_fn=p_component_fn,
            p_fastener_fn=p_fastener_fn,
            tools_unc_by_part=tools_unc,
            rng=rng,
            pop=args.pop,
            gens=args.gens,
            sims=args.sims,
            force_scale=args.force_scale,
            mutation_sigma=args.mutation_sigma,
            elite_frac=args.elite_frac,
            dynamic_mc=args.dynamic_mc,
            sims_coarse=args.sims_coarse,
            fine_top_frac=args.fine_top_frac,
            task_difficulty_by_code=task_difficulty_by_code,
        )

        # Print the induced greedy sequence from best weights
        def policy_weights_choose(available, removed_set, wear_state):
            best = None
            best_w = -1e18
            for v in available:
                w = best_weights.get(v, 0.0)
                if w > best_w:
                    best_w = w
                    best = v
            return best

        removed: Set[str] = set()
        remaining = set(nodes_to_consider)
        seq = []
        while args.goal not in removed:
            available = topo_available(removed, preds, remaining)
            if not available:
                break
            v = policy_weights_choose(available, removed, 0.0)
            removed.add(v)
            seq.append(v)

        print("\n[GA] Best priority-induced sequence until goal:")
        print(" -> ".join(seq))
        print("[GA] Tasks:")
        print(" -> ".join(f"{v}({task_for_node(v)})" for v in seq))
        print(f"[GA] Best mean_cost during GA fitness: {best_cost:.4f}")


if __name__ == "__main__":
    main()

