from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Dict, List, Set, Tuple

_ROOT = Path(__file__).resolve().parent


def parse_edges_and_partnames(html_path: str) -> Tuple[Set[str], List[Dict[str, str]], Dict[str, str], Dict[str, int]]:
    with open(html_path, "r", encoding="utf-8") as f:
        txt = f.read()

    # edges: ['F1','F2'] etc
    start = txt.find("const edges")
    if start < 0:
        raise RuntimeError("Could not find const edges in graph html")
    end = txt.find("].map", start)
    if end < 0:
        raise RuntimeError("Could not find edges end marker in graph html")

    block = txt[start:end]
    pair_pat = re.compile(r"\[\s*'([^']+)'\s*,\s*'([^']+)'\s*\]")
    pairs = pair_pat.findall(block)
    if not pairs:
        raise RuntimeError("No edges parsed from graph html")

    edges = [{"from": a, "to": b} for a, b in pairs]

    nodes: Set[str] = set()
    for e in edges:
        nodes.add(e["from"])
        nodes.add(e["to"])

    # disassemblyStep (removal step mapping)
    dispatch: Dict[str, int] = {}
    m = re.search(
        r"const\s+disassemblyStep\s*=\s*\{([^}]*)\}\s*;", txt, flags=re.S)
    if m:
        body = m.group(1)
        for k, v in re.findall(r"([A-Za-z0-9]+)\s*:\s*(\d+)", body):
            dispatch[k] = int(v)

    # partNames
    partNames: Dict[str, str] = {}
    m2 = re.search(r"const\s+partNames\s*=\s*\{([^}]*)\}\s*;", txt, flags=re.S)
    if m2:
        b2 = m2.group(1)
        # lines like: F1: 'Screw ...',
        for k, v in re.findall(r"(F[0-9A-Z]+|C[0-9A-Z]+)\s*:\s*'([^']*)'", b2):
            partNames[k] = v

    return nodes, edges, partNames, dispatch


def read_xlsx_uncertainties(xlsx_path: str):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def _norm_header(x):
        return re.sub(r"\s+", " ", str(x).strip().lower()) if x is not None else ""

    def _build_col_map(ws, header_row: int):
        headers = [ws.cell(header_row, c).value for c in range(
            1, ws.max_column + 1)]
        m = {}
        for idx, h in enumerate(headers, start=1):
            key = _norm_header(h)
            if not key:
                continue
            # Keep first occurrence
            m.setdefault(key, idx)
        return m

    def _find_task_table(ws):
        # Look for a row containing "Task code" and "Difficulty"
        for r in range(1, min(500, ws.max_row + 1)):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if a is None or b is None:
                continue
            if "task code" in str(a).lower() and "difficulty" in str(b).lower():
                return r
        return None

    ws_comp = wb["components"]
    comp_row1 = _build_col_map(ws_comp, header_row=1)
    comp_row2 = _build_col_map(ws_comp, header_row=2)
    components_unc_by_part: Dict[int, Dict[str, float]] = {}
    component_task_by_part: Dict[int, str] = {}

    part_col = comp_row1.get("part number", 1)
    task_col = comp_row1.get("task", 4)
    temp_col = comp_row2.get("temperature", 5)
    corr_col = comp_row2.get("corrosion level", 6)

    task_table_header_row = _find_task_table(ws_comp)
    stop_row = (task_table_header_row -
                1) if task_table_header_row else (ws_comp.max_row + 1)

    for r in range(3, stop_row + 1):
        row_part = ws_comp.cell(r, part_col).value
        if row_part is None:
            continue
        try:
            part = int(float(row_part))
        except Exception:
            continue
        temp = ws_comp.cell(r, temp_col).value
        corr = ws_comp.cell(r, corr_col).value
        task = ws_comp.cell(r, task_col).value
        if temp is None or corr is None:
            continue
        components_unc_by_part[part] = {
            "temperature": float(temp), "corrosion_level": float(corr)}
        if task is not None and str(task).strip():
            component_task_by_part[part] = str(task).strip()

    ws_fast = wb["fasteners"]
    fast_row1 = _build_col_map(ws_fast, header_row=1)
    fast_row2 = _build_col_map(ws_fast, header_row=2)
    tools_unc_by_part: Dict[int, Dict[str, float]] = {}
    fastener_task_by_part: Dict[int, str] = {}
    fastener_tool_code_by_part: Dict[int, str] = {}

    part_col_f = fast_row1.get("part number", 1)
    task_col_f = fast_row1.get("task", 4)

    tool_code_col_f = fast_row1.get("tool code") or fast_row1.get(
        "tool") or fast_row1.get("tool/alttools")

    success_col = fast_row2.get("fasteners success rate") or fast_row2.get(
        "tool success rate") or 5
    wear_col = fast_row2.get(
        "fasteners wear") or fast_row2.get("tool wear") or 6
    temp_col_f = fast_row2.get("temperature") or 7
    bolt_col = fast_row2.get("bolt seizure probability") or 9
    jam_col = fast_row2.get("bearing jamming probability") or 10
    force_col = fast_row2.get("removal force (nm)") or 11

    for r in range(3, ws_fast.max_row + 1):
        row_part = ws_fast.cell(r, part_col_f).value
        if row_part is None:
            continue
        try:
            part = int(float(row_part))
        except Exception:
            continue
        tool_success_rate = ws_fast.cell(r, success_col).value
        tool_wear = ws_fast.cell(r, wear_col).value
        tool_temperature = ws_fast.cell(r, temp_col_f).value
        bolt_seiz = ws_fast.cell(r, bolt_col).value
        bearing_jam = ws_fast.cell(r, jam_col).value
        removal_force = ws_fast.cell(r, force_col).value
        task = ws_fast.cell(r, task_col_f).value if task_col_f else None
        tool_code = ws_fast.cell(
            r, tool_code_col_f).value if tool_code_col_f else None

        if tool_success_rate is None or tool_wear is None:
            continue

        tools_unc_by_part[part] = {
            "tool_success_rate": float(tool_success_rate),
            "tool_wear": float(tool_wear),
            "temperature": float(tool_temperature) if tool_temperature is not None else 0.0,
            "bolt_seizure_probability": float(bolt_seiz) if bolt_seiz is not None else 0.0,
            "bearing_jamming_probability": float(bearing_jam) if bearing_jam is not None else 0.0,
            "removal_force_Nm": float(removal_force) if removal_force is not None else 0.0,
        }
        if task is not None and str(task).strip():
            fastener_task_by_part[part] = str(task).strip()
        if tool_code is not None and str(tool_code).strip():
            fastener_tool_code_by_part[part] = str(tool_code).strip()

    task_difficulty_by_code: Dict[str, float] = {}
    if task_table_header_row:
        for r in range(task_table_header_row + 1, min(task_table_header_row + 50, ws_comp.max_row + 1)):
            code_cell = ws_comp.cell(r, 1).value
            diff_cell = ws_comp.cell(r, 2).value
            if code_cell is None or diff_cell is None:
                continue
            raw = str(code_cell).strip()
            if not raw:
                continue
            code = raw.split("-", 1)[0].strip()
            try:
                diff = float(diff_cell)
            except Exception:
                continue
            if code:
                task_difficulty_by_code[code] = diff

    # component_removal: component node -> BOM part code
    ws5 = wb["component_removal"]
    comp_node_to_part: Dict[str, int] = {}
    for row in ws5.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        node = str(row[0])
        bom = row[1]
        if bom is None:
            continue
        comp_node_to_part[node] = int(float(bom))

    # fasteners_removal: fastener node -> BOM part code
    ws6 = wb["fasteners_removal"]
    fast_node_to_part: Dict[str, int] = {}
    for row in ws6.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        node = str(row[0])
        bom = row[1]
        if bom is None:
            continue
        fast_node_to_part[node] = int(float(bom))

    return (
        components_unc_by_part,
        tools_unc_by_part,
        comp_node_to_part,
        fast_node_to_part,
        component_task_by_part,
        fastener_task_by_part,
        task_difficulty_by_code,
        fastener_tool_code_by_part,
    )


def main():
    html_path = str(_ROOT.parent / "script" / "gearbox_kg.html")
    xlsx_path = str(_ROOT / "gearbox.xlsx")
    out_path = str(_ROOT / "build_data.json")

    nodes, edges, partNames, dispatch = parse_edges_and_partnames(html_path)
    (
        components_unc_by_part,
        tools_unc_by_part,
        comp_node_to_part,
        fast_node_to_part,
        component_task_by_part,
        fastener_task_by_part,
        task_difficulty_by_code,
        fastener_tool_code_by_part,
    ) = read_xlsx_uncertainties(xlsx_path)

    node_records: List[Dict] = []
    component_temps: List[float] = []

    for nid in sorted(nodes):
        kind = "F" if nid.startswith("F") else "C"
        part_code = None
        if kind == "C":
            part_code = comp_node_to_part.get(nid)
        else:
            part_code = fast_node_to_part.get(nid)

        label = partNames.get(nid, nid)

        task_code = None
        tool_code = None
        if kind == "C" and part_code is not None:
            task_code = component_task_by_part.get(part_code)
        if kind == "F" and part_code is not None:
            task_code = fastener_task_by_part.get(part_code)
            tool_code = fastener_tool_code_by_part.get(part_code)

        task_code_norm = task_code.strip() if isinstance(task_code, str) else None
        task_difficulty = task_difficulty_by_code.get(
            task_code_norm) if task_code_norm else None

        rec = {
            "id": nid,
            "kind": kind,
            "partCode": part_code,
            "label": label,
            "step": dispatch.get(nid),
            "taskCode": task_code_norm,
            "taskDifficulty": task_difficulty,
            "toolCode": tool_code.strip() if isinstance(tool_code, str) and tool_code.strip() else None,
            "component": None,
            "fastener": None,
            "tool": None,
        }
        if kind == "C" and part_code is not None:
            comp = components_unc_by_part.get(part_code)
            if comp is not None:
                rec["component"] = comp
                component_temps.append(comp["temperature"])
        if kind == "F" and part_code is not None:
            tool = tools_unc_by_part.get(part_code)
            if tool is not None:
                # Prefer `fastener` naming in consumers; keep `tool` for existing code.
                rec["fastener"] = tool
                rec["tool"] = tool

        node_records.append(rec)

    comp_temp_min = min(component_temps) if component_temps else 0.0
    comp_temp_max = max(component_temps) if component_temps else 1.0

    out = {
        "nodes": node_records,
        "edges": edges,
        "goal": "C1",
        "componentTempMin": comp_temp_min,
        "componentTempMax": comp_temp_max,
        "disassemblyStep": dispatch,
        "taskDifficultyByCode": task_difficulty_by_code,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2)

    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
